import yfinance as yf
import json
from pathlib import Path
from dotenv import load_dotenv
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tkinter as tk
from tkinter import ttk, messagebox
import re
import time
from datetime import datetime

load_dotenv()

DATA_FILE = Path("dividend_data.json")


class DataFileError(Exception):
    """Raised when dividend_data.json exists but cannot be read."""


def load_dividend_data(path=DATA_FILE):
    """Return the saved ticker records, or [] if the file doesn't exist yet.

    Raises DataFileError if the file exists but can't be parsed, so callers can
    warn the user instead of silently overwriting data that may be recoverable.
    """
    if not path.is_file():
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise DataFileError(f"{path} is not valid JSON, so it was left untouched: {e}") from e
    except OSError as e:
        raise DataFileError(f"Could not read {path}: {e}") from e

    if not isinstance(data, list):
        raise DataFileError(f"{path} should contain a list of tickers, found {type(data).__name__}.")
    return data


def save_dividend_data(data, path=DATA_FILE):
    """Write the ticker records, replacing the file only once it's fully written.

    Writing straight to the real path truncates it the moment it's opened, so an
    interrupted write leaves a half-file that fails to parse on next launch.
    """
    tmp = path.with_name(path.name + ".tmp")
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        os.replace(tmp, path)
    except OSError as e:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        raise DataFileError(f"Could not save {path}: {e}") from e

class ValidationUtils:
    """Utility class for input validation and error handling"""

    @staticmethod
    def validate_ticker_symbol(symbol):
        """Validate ticker symbol format"""
        if not symbol or not isinstance(symbol, str):
            return False, "Ticker symbol cannot be empty"

        symbol = symbol.strip().upper()

        # Basic format validation
        if len(symbol) < 1 or len(symbol) > 10:
            return False, "Ticker symbol must be 1-10 characters long"

        # Colons are matched here only so the ":CA" check below can return a more
        # useful message than this generic one.
        if not re.match(r'^[A-Z0-9\.:]+$', symbol):
            return False, "Ticker symbol can only contain letters, numbers, and dots"

        # Check for common invalid patterns
        if symbol.startswith('.') or symbol.endswith('.'):
            return False, "Ticker symbol cannot start or end with a dot"

        if '..' in symbol:
            return False, "Ticker symbol cannot contain consecutive dots"

        # Yahoo Finance identifies Canadian listings by exchange suffix (.TO, .V,
        # .CN), not the ":CA" convention other providers use. ":CA" symbols come
        # back empty rather than as an error, so reject them with guidance here.
        if ':' in symbol:
            return False, ("Use the exchange suffix instead of a colon, "
                           "e.g. 'RY.TO' (Toronto), '.V' (TSX Venture) or '.CN' (CSE)")

        return True, symbol

    @staticmethod
    def is_duplicate_ticker(symbol, existing_tickers):
        """Check if ticker already exists in portfolio"""
        symbol = symbol.strip().upper()
        for ticker in existing_tickers:
            if hasattr(ticker, 'symbol') and ticker.symbol == symbol:
                return True
        return False

class App():
    def __init__(self, root):
        self.root = root
        self.root.title("Dividend Tracker")
        self.root.geometry("600x600")

        self.style = ttk.Style()
        self.style.theme_use("clam")

        self.dataManager = DividendDataManager()
        self.setup_ui()

        # Surface a bad data file in the UI rather than dying before it exists.
        if self.dataManager.load_error:
            messagebox.showwarning(
                "Could Not Load Portfolio",
                f"{self.dataManager.load_error}\n\n"
                "Started with an empty list. Your file has not been changed, so "
                "you can fix or move it and restart."
            )



    def setup_ui(self):
        # Frame for ticker list
        list_frame = ttk.LabelFrame(self.root, text="Current Stock Tickers")
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.ticker_listbox = tk.Listbox(list_frame, height=10, font=("Segoe UI", 12))
        self.ticker_listbox.pack(side="left", fill="both", expand=True, padx=(10,0), pady=10)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.ticker_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.ticker_listbox.config(yscrollcommand=scrollbar.set)
        # Populate listbox
        for ticker in self.dataManager.tickers:
           self.ticker_listbox.insert(tk.END, ticker.symbol)

        # Frame for adding/removing tickers
        control_frame = ttk.Frame(self.root)
        control_frame.pack(fill="x", padx=10, pady=10)

        self.ticker_entry = ttk.Entry(control_frame, font=("Segoe UI", 12))
        self.ticker_entry.pack(side="left", fill="x", expand=True, padx=(0,10))

        add_button = ttk.Button(control_frame, text="Add Ticker", command=self.add_ticker)
        add_button.pack(side="left", padx=(0,10))

        remove_button = ttk.Button(control_frame, text="Remove Selected", command=self.remove_selected_ticker)
        remove_button.pack(side="left")

        # Export/update buttons at bottom
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(fill="x", padx=10, pady=10)

        update_button = ttk.Button(bottom_frame, text="Build/Rebuild Excel", command=self.build_excel)
        update_button.pack(side="left", padx=(0,10))

        export_button = ttk.Button(bottom_frame, text="Open Excel", command=self.open_excel)
        export_button.pack(side="left")



    def remove_selected_ticker(self):
        """Remove the selected ticker, updating the list only once the save succeeds."""
        selected = self.ticker_listbox.curselection()
        if not selected:
            messagebox.showinfo("No Selection", "Select a ticker to remove first.")
            return

        index = selected[0]
        symbol = self.ticker_listbox.get(index)

        # Save before touching the listbox
        success, error = self.dataManager.remove_ticker(symbol)
        if not success:
            messagebox.showerror("Remove Failed", f"Could not remove {symbol}.\n\n{error}")
            return

        self.ticker_listbox.delete(index)

    def add_ticker(self):
        """Add a new ticker with validation and error handling"""
        try:
            new_ticker = self.ticker_entry.get().strip()

            # Validate ticker symbol
            is_valid, result = ValidationUtils.validate_ticker_symbol(new_ticker)
            if not is_valid:
                messagebox.showerror("Invalid Ticker", f"Error: {result}")
                return

            new_ticker = result  # Use the validated/cleaned symbol

            # Check for duplicates
            if ValidationUtils.is_duplicate_ticker(new_ticker, self.dataManager.tickers):
                messagebox.showwarning("Duplicate Ticker", f"'{new_ticker}' is already in your portfolio")
                return

            # Show loading state
            self.show_loading_state()

            # Add ticker with error handling
            success, error = self.dataManager.add_ticker(new_ticker)

            if success:
                self.ticker_listbox.insert(tk.END, new_ticker)
                messagebox.showinfo("Success", f"Successfully added {new_ticker} to your portfolio")
            else:
                messagebox.showerror("Error", f"Failed to add {new_ticker}.\n\n{error}")

        except Exception as e:
            messagebox.showerror("Unexpected Error", f"An unexpected error occurred: {str(e)}")
        finally:
            # Re-enable first: Tk silently ignores delete() on a disabled entry,
            # which would leave the previous symbol sitting in the box.
            self.hide_loading_state()
            self.ticker_entry.delete(0, tk.END)

    def show_loading_state(self):
        """Show loading indicator during API calls"""
        self.ticker_entry.config(state='disabled')

    def hide_loading_state(self):
        """Hide loading indicator after API calls"""
        self.ticker_entry.config(state='normal')

    def build_excel(self):
        """Build Excel file with comprehensive error handling"""
        try:
            if not DATA_FILE.is_file():
                messagebox.showerror("File Not Found", f"{DATA_FILE} not found. Please add some tickers first.")
                return

            try:
                dividend_data = load_dividend_data()
            except DataFileError as e:
                messagebox.showerror("Data Error", str(e))
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Dividends"

            # Headers - expanded to include all Alpha Vantage fields
            headers = ["Ex-Date", "Declaration Date", "Record Date", "Payment Date", "Ticker", "Currency", "Dividend"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")

            for col_num, title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

            # Column widths
            ws.column_dimensions["A"].width = 15  # Ex-Date
            ws.column_dimensions["B"].width = 15  # Declaration Date
            ws.column_dimensions["C"].width = 15  # Record Date
            ws.column_dimensions["D"].width = 15  # Payment Date
            ws.column_dimensions["E"].width = 12  # Ticker
            ws.column_dimensions["F"].width = 10  # Currency
            ws.column_dimensions["G"].width = 12  # Dividend

            # Fill rows
            row_num = 2
            for entry in dividend_data:
                ticker = entry.get("ticker")
                currency = entry.get("currency", "USD")
                for div in entry.get("dividends", []):
                    # Handle both TSX (ex_date) and US (ex_dividend_date) formats
                    ex_date = div.get("ex_date") or div.get("ex_dividend_date")
                    declaration_date = div.get("declaration_date")
                    record_date = div.get("record_date")
                    payment_date = div.get("payment_date")
                    amount = div.get("amount")

                    # Convert amount to float for proper Excel number formatting
                    try:
                        amount = float(amount) if amount else 0
                    except (ValueError, TypeError):
                        amount = 0

                    # Convert "None" strings to empty cells for better Excel display
                    if declaration_date == "None":
                        declaration_date = ""
                    if record_date == "None":
                        record_date = ""
                    if payment_date == "None":
                        payment_date = ""

                    ws.cell(row=row_num, column=1, value=ex_date)
                    ws.cell(row=row_num, column=2, value=declaration_date)
                    ws.cell(row=row_num, column=3, value=record_date)
                    ws.cell(row=row_num, column=4, value=payment_date)
                    ws.cell(row=row_num, column=5, value=ticker)
                    ws.cell(row=row_num, column=6, value=currency)
                    ws.cell(row=row_num, column=7, value=amount)
                    row_num += 1

            # Save workbook
            output_path = "dividends-sheet.xlsx"
            try:
                wb.save(output_path)
                messagebox.showinfo("Success", f"Excel file saved successfully!\nLocation: {output_path}")
                return output_path
            except PermissionError:
                messagebox.showerror("Permission Error",
                    "Cannot save Excel file. Please close the file if it's open in Excel and try again.")
                return None
            except Exception as e:
                messagebox.showerror("Save Error", f"Error saving Excel file: {str(e)}")
                return None

        except Exception as e:
            messagebox.showerror("Excel Export Error", f"An error occurred while building the Excel file:\n{str(e)}")
            return None


    def open_excel(self):
        """Open the Excel file with the system's default program"""
        import subprocess
        import sys

        excel_path = "dividends-sheet.xlsx"

        # Check if file exists
        if not os.path.exists(excel_path):
            messagebox.showwarning("File Not Found",
                "Excel file not found. Please click 'Rebuild Excel' first to generate the file.")
            return

        try:
            # Open file with system default program
            if sys.platform == "win32":
                os.startfile(excel_path)
            elif sys.platform == "darwin":  # macOS
                subprocess.run(["open", excel_path])
            else:  # Linux and others
                subprocess.run(["xdg-open", excel_path])
            messagebox.showinfo("Success", f"Opening {excel_path}...")
        except FileNotFoundError:
            messagebox.showerror("Application Not Found",
                "No application found to open Excel files. Please install Excel or a compatible spreadsheet application.")
        except Exception as e:
            messagebox.showerror("Error", f"Error opening Excel file: {str(e)}")


class DividendDataManager:
    # Gather all ticker data necessary for excel and json...
    def __init__(self):
        self.tickers = []
        self.load_error = None

        try:
            data = load_dividend_data()
        except DataFileError as e:
            # An unreadable file must not stop the window from opening; App shows
            # this once the UI exists, instead of dying with a console traceback.
            self.load_error = str(e)
            return

        for item in data:
            symbol = item.get("ticker")
            if not symbol:
                continue  # skip a malformed record rather than fail to start
            self.tickers.append(StockTicker(symbol))

    def add_ticker(self, symbol):
        """Fetch and save a ticker. Returns (success, error_message)."""
        try:
            ticker = StockTicker(symbol, True)
        except Exception as e:
            return False, f"Could not fetch data for {symbol}: {e}"

        if not ticker.fetch_ok:
            return False, ticker.fetch_error

        try:
            self.add_to_json(ticker)
        except DataFileError as e:
            return False, str(e)

        # Only track it in memory once it's safely on disk.
        self.tickers.append(ticker)
        return True, None

    def add_to_json(self, ticker):
        data = load_dividend_data()
        data.append(ticker.data)
        save_dividend_data(data)

    def remove_ticker(self, symbol):
        """Remove a ticker from disk, then memory. Returns (success, error_message)."""
        try:
            self.remove_from_json(symbol)
        except DataFileError as e:
            return False, str(e)

        self.tickers = [t for t in self.tickers if t.symbol != symbol]
        return True, None

    def remove_from_json(self, symbol):
        data = load_dividend_data()
        data = [entry for entry in data if entry.get("ticker") != symbol]
        save_dividend_data(data)




class StockTicker:
    def __init__(self, symbol, new=False):
        self.symbol = symbol.strip().upper()
        # A valid ticker that simply paid nothing in the lookback window still
        # counts as a successful fetch, so success is tracked separately from
        # whether any dividends came back.
        self.fetch_ok = True
        self.fetch_error = None
        self.data = {
            "ticker": self.symbol,
            "currency": "USD",
            "dividends": []
        }
        if new == True:
            if self.symbol.endswith((".TO", ".V", ".CN")):
                self.fetch_tsx_dividends()
            else:
                self.fetch_dividends()

    def _fail(self, message):
        """Record why a fetch failed so the caller can show a real reason."""
        self.fetch_ok = False
        self.fetch_error = message
        self.data["dividends"] = []

    def _symbol_exists(self):
        """Whether Yahoo recognises this symbol.

        An empty dividend series means either an unknown symbol or a stock that
        pays nothing, and yfinance can't tell them apart, so fall back to asking
        for recent price history.
        """
        try:
            return not self.ticker.history(period="5d").empty
        except Exception:
            return False

    def fetch_tsx_dividends(self):
        """Fetch Canadian listing dividends via yfinance."""
        print("Dividends for TSX stocks are being fetched!")
        self.data["currency"] = "CAD"

        try:
            self.ticker = yf.Ticker(self.symbol)
            div_series = self.ticker.get_dividends(period="3mo")
        except Exception as e:
            self._fail(f"Could not reach Yahoo Finance for {self.symbol}: {e}")
            return

        # yfinance returns an empty Series for an unknown symbol rather than
        # raising, so an empty result needs a second look before trusting it.
        if div_series is None or div_series.empty:
            if not self._symbol_exists():
                self._fail(f"No data found for {self.symbol}. Check the symbol, "
                           f"its exchange suffix, and your connection.")
            else:
                self.data["dividends"] = []
            return

        dividends = []
        for date, amount in div_series.items():
            try:
                dividends.append({
                    "ex_dividend_date": date.strftime("%Y-%m-%d"),
                    "amount": float(amount),
                })
            except (ValueError, TypeError):
                continue  # drop an unparseable row rather than lose the whole fetch

        self.data["dividends"] = dividends
        print(f"Successfully fetched {len(dividends)} dividend records for {self.symbol}")

    def fetch_dividends(self):
        """Fetch US stock dividends"""
        from datetime import timedelta
        print("Dividends for US stocks are being fetched!")

        # Get Alpha Vantage API key from environment
        api_key = os.getenv("ALPHA_VANTAGE_API_KEY")
        if not api_key:
            self._fail("ALPHA_VANTAGE_API_KEY not found. Add it to your .env file.")
            return

        # Alpha Vantage API URL for dividend data
        url = f'https://www.alphavantage.co/query?function=DIVIDENDS&symbol={self.symbol}&apikey={api_key}'

        # Retry logic for network failures
        max_retries = 3
        retry_delay = 1  # seconds

        for attempt in range(max_retries):
            try:
                print(f"Fetching data for {self.symbol} (attempt {attempt + 1}/{max_retries})")

                # Make API request with timeout
                response = requests.get(url, timeout=30)
                response.raise_for_status()
                data = response.json()

                # Check for API errors
                if "Error Message" in data:
                    self._fail(f"Alpha Vantage rejected '{self.symbol}': {data['Error Message']}")
                    return

                # Throttling comes back as HTTP 200 with an explanatory key, so
                # raise_for_status() won't catch it. Alpha Vantage has used both
                # "Note" and "Information" for this; checking only one means the
                # backoff below never runs and the result looks like "no data".
                throttle = data.get("Note") or data.get("Information")
                if throttle:
                    print(f"API Rate Limit for {self.symbol}: {throttle}")
                    if attempt < max_retries - 1:
                        print(f"Waiting {retry_delay * 2} seconds before retry...")
                        time.sleep(retry_delay * 2)
                        retry_delay *= 2  # Exponential backoff
                        continue
                    else:
                        self._fail(f"Alpha Vantage rate limit reached for {self.symbol}: {throttle}")
                        return

                if "data" not in data:
                    self._fail(f"Unexpected response from Alpha Vantage for {self.symbol}.")
                    return

                # Filter to last 6 months
                cutoff_date = datetime.now() - timedelta(days=180)

                filtered_dividends = []
                for dividend in data["data"]:
                    ex_date_str = dividend.get("ex_dividend_date")
                    if ex_date_str and ex_date_str != "None":
                        try:
                            ex_date = datetime.strptime(ex_date_str, "%Y-%m-%d")
                            if ex_date >= cutoff_date:
                                filtered_dividends.append(dividend)
                        except ValueError:
                            continue

                # Sort by date (newest first)
                filtered_dividends.sort(key=lambda x: x.get("ex_dividend_date", ""), reverse=True)

                self.data["dividends"] = filtered_dividends
                print(f"Successfully fetched {len(filtered_dividends)} dividend records for {self.symbol}")
                return  # Success, exit retry loop

            except requests.exceptions.Timeout:
                print(f"Timeout error for {self.symbol} (attempt {attempt + 1})")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    self._fail(f"Alpha Vantage timed out for {self.symbol} after {max_retries} attempts.")

            except requests.exceptions.ConnectionError:
                print(f"Connection error for {self.symbol} (attempt {attempt + 1})")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    self._fail(f"Could not connect to Alpha Vantage for {self.symbol}. "
                               f"Check your internet connection.")

            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code == 429:  # Rate limit
                    print(f"Rate limit exceeded for {self.symbol}")
                    if attempt < max_retries - 1:
                        wait_time = retry_delay * 5  # Longer wait for rate limits
                        print(f"Waiting {wait_time} seconds before retry...")
                        time.sleep(wait_time)
                        retry_delay *= 2
                    else:
                        self._fail(f"Alpha Vantage rate limit reached for {self.symbol}.")
                else:
                    self._fail(f"Alpha Vantage returned an HTTP error for {self.symbol}: {e}")
                    return

            except Exception as e:
                self._fail(f"Unexpected error fetching {self.symbol}: {e}")
                return




if __name__ == "__main__":
    import ctypes



    # Optional: make app DPI aware on Windows
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass



    root = tk.Tk()  # Create the main Tkinter window
    app = App(root)  # Create your app instance



    root.mainloop()  # Start Tkinter's main loop
